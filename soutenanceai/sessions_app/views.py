"""Vues du dashboard professeur : classes, sessions, planning, auto-planification."""
import random
import unicodedata
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Max
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from accounts.decorators import professeur_required
from accounts.models import User

from .constants import CRITERES_PREDEFINIS
from .forms import (
    ClasseForm, ImportGroupesForm, PassageForm, PlanifierAutoForm, PlanifierPassageForm,
    RejoindreClasseForm, RejoindreSessionForm, SessionForm,
)
from .importation import groupes_to_json, matcher_groupes, parse_fichier
from .models import Classe, CritereNotation, PassageEtudiant, Session


# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm_username(prenom, nom):
    """Génère username prenom.nom en ASCII minuscules, gère les doublons."""
    def _clean(s):
        return ''.join(
            c for c in unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
            if c.isalnum()
        ).lower()
    base = f"{_clean(prenom)}.{_clean(nom)}"
    if base and not User.objects.filter(username=base).exists():
        return base
    i = 2
    while User.objects.filter(username=f"{base}_{i}").exists():
        i += 1
    return f"{base}_{i}" if base else f"etudiant_{i}"


@professeur_required
def dashboard_professeur(request):
    classes = Classe.objects.filter(professeur=request.user).prefetch_related('sessions').order_by('-date_creation')
    return render(request, 'sessions_app/dashboard.html', {'classes': classes})


# ── CRUD Classe ───────────────────────────────────────────────────────────────

@professeur_required
@require_http_methods(['GET', 'POST'])
def creer_classe(request):
    form = ClasseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        classe = form.save(commit=False)
        classe.professeur = request.user
        classe.save()
        messages.success(request, f'Classe « {classe.nom} » créée !')
        return redirect('sessions_app:code_classe', classe_id=classe.id)
    return render(request, 'sessions_app/form_classe.html', {'form': form, 'classe': None})


@professeur_required
def detail_classe(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user)
    sessions = classe.sessions.prefetch_related('passages').order_by('-date_creation')
    return render(request, 'sessions_app/detail_classe.html', {
        'classe': classe,
        'sessions': sessions,
    })


@professeur_required
@require_http_methods(['GET', 'POST'])
def editer_classe(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user)
    form = ClasseForm(request.POST or None, instance=classe)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Classe mise à jour.')
        return redirect('sessions_app:detail_classe', classe_id=classe.id)
    return render(request, 'sessions_app/form_classe.html', {'form': form, 'classe': classe})


@professeur_required
@require_http_methods(['POST'])
def supprimer_classe(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user)
    nom = classe.nom
    classe.delete()
    messages.success(request, f'Classe « {nom} » supprimée.')
    return redirect('sessions_app:dashboard_professeur')


@professeur_required
def code_classe(request, classe_id):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user)
    join_url = request.build_absolute_uri(f'/classes/rejoindre/{classe.code_acces}/')
    return render(request, 'sessions_app/code_classe.html', {
        'classe': classe,
        'join_url': join_url,
    })


@professeur_required
@require_http_methods(['POST'])
def retirer_inscrit_classe(request, classe_id, user_id):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user)
    user = get_object_or_404(User, pk=user_id)
    classe.inscrits.remove(user)
    messages.success(request, f'{user.get_full_name() or user.username} retiré(e) de la classe.')
    return redirect('sessions_app:detail_classe', classe_id=classe.id)


# ── Rejoindre une classe (public — étudiant) ──────────────────────────────────

def rejoindre_classe(request, code):
    """Page publique : un étudiant rejoint une Classe via son code."""
    classe = get_object_or_404(Classe, code_acces=code.upper())

    if request.user.is_authenticated:
        if request.user.is_etudiant:
            classe.inscrits.add(request.user)
            messages.success(request, f'Vous avez rejoint la classe « {classe.nom} ».')
            return redirect('presentation:dashboard_etudiant')
        messages.info(request, 'Seuls les étudiants peuvent rejoindre une classe.')
        return redirect('accounts:dashboard_redirect')

    form = RejoindreClasseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        data = form.cleaned_data
        if data.get('email') and User.objects.filter(email=data['email']).exists():
            form.add_error('email', 'Cet email est déjà utilisé. Connectez-vous plutôt.')
        else:
            user = User.objects.create_user(
                username=data['username'],
                first_name=data['prenom'],
                last_name=data['nom'],
                email=data.get('email', ''),
                password=data['password'],
                role=User.ROLE_ETUDIANT,
                cree_par=classe.professeur,
            )
            classe.inscrits.add(user)
            auth_login(request, user)
            messages.success(
                request,
                f'Bienvenue {user.get_full_name()} ! Vous avez rejoint « {classe.nom} ».',
            )
            return redirect('presentation:dashboard_etudiant')

    return render(request, 'sessions_app/rejoindre_classe.html', {
        'classe': classe,
        'form': form,
        'code': code.upper(),
    })


@professeur_required
@require_http_methods(['GET', 'POST'])
def creer_session(request, classe_id=None):
    classe = get_object_or_404(Classe, pk=classe_id, professeur=request.user) if classe_id else None
    return _creer_ou_editer_session(request, session=None, classe=classe)


@professeur_required
@require_http_methods(['GET', 'POST'])
def editer_session(request, session_id):
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    return _creer_ou_editer_session(request, session=session)


def _creer_ou_editer_session(request, session, classe=None):
    form = SessionForm(request.POST or None, instance=session)

    # Critères : on récupère ceux pré-cochés (POST) ou existants (édition)
    if session:
        criteres_existants = list(session.criteres.all().order_by('ordre'))
        noms_predef_coches = [
            c.nom for c in criteres_existants if not c.est_personnalise
        ]
        criteres_perso = [c for c in criteres_existants if c.est_personnalise]
        coefs_existants = {c.nom: c.coefficient for c in criteres_existants}
    else:
        noms_predef_coches = []
        criteres_perso = []
        coefs_existants = {}

    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            session_obj = form.save(commit=False)
            session_obj.professeur = request.user
            if classe and not session_obj.pk:
                session_obj.classe = classe
            session_obj.save()

            # Critères prédéfinis cochés
            rapport_obligatoire = form.cleaned_data.get('rapport_obligatoire', False)
            session_obj.criteres.all().delete()
            ordre = 0
            for nom, _desc in CRITERES_PREDEFINIS:
                # Contrainte serveur : "Rapport écrit" ignoré si rapport non obligatoire
                if nom == 'Rapport écrit' and not rapport_obligatoire:
                    continue
                if request.POST.get(f'critere_predef_{nom}'):
                    coef = float(request.POST.get(f'coef_predef_{nom}', '1') or 1)
                    CritereNotation.objects.create(
                        session=session_obj, nom=nom,
                        coefficient=coef, est_personnalise=False, ordre=ordre,
                    )
                    ordre += 1
            # Critères personnalisés (champs nommés critere_perso_N + coef_perso_N)
            perso_idx = 0
            while True:
                nom = request.POST.get(f'critere_perso_{perso_idx}', '').strip()
                if not nom and perso_idx > 50:
                    break
                if nom:
                    coef = float(request.POST.get(f'coef_perso_{perso_idx}', '1') or 1)
                    CritereNotation.objects.create(
                        session=session_obj, nom=nom,
                        coefficient=coef, est_personnalise=True, ordre=ordre,
                    )
                    ordre += 1
                perso_idx += 1
                if perso_idx > 100:
                    break

        messages.success(request, f'Session "{session_obj.titre}" enregistrée.')
        return redirect('sessions_app:detail_session', session_id=session_obj.id)

    return render(request, 'sessions_app/form_session.html', {
        'form': form,
        'session': session,
        'classe': classe or (session.classe if session else None),
        'criteres_predefinis': CRITERES_PREDEFINIS,
        'noms_predef_coches': noms_predef_coches,
        'criteres_perso': criteres_perso,
        'coefs_existants': coefs_existants,
    })


@professeur_required
def detail_session(request, session_id):
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    passages = session.passages.prefetch_related('etudiants').order_by('ordre_passage')
    return render(request, 'sessions_app/detail_session.html', {
        'session': session,
        'passages': passages,
        'criteres': session.criteres.all(),
    })


@professeur_required
@require_http_methods(['POST'])
def supprimer_session(request, session_id):
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    session.delete()
    messages.success(request, 'Session supprimée.')
    return redirect('sessions_app:dashboard_professeur')


@professeur_required
@require_http_methods(['GET', 'POST'])
def editer_passage(request, passage_id):
    """Édition d'un passage existant (heure, ordre, étudiants).
    Accessible via le bouton crayon dans le tableau de planification.
    La création de passage se fait directement dans planifier_passages (formulaire inline).
    """
    passage = get_object_or_404(
        PassageEtudiant, pk=passage_id, session__professeur=request.user,
    )
    form = PassageForm(request.POST or None, instance=passage, professeur=request.user)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Passage mis a jour.')
        # Retour vers la page de planification
        return redirect('sessions_app:planifier_passages', session_id=passage.session.id)
    return render(request, 'sessions_app/form_passage.html', {
        'form': form, 'session': passage.session, 'passage': passage,
    })


@professeur_required
@require_http_methods(['POST'])
def supprimer_passage(request, passage_id):
    passage = get_object_or_404(
        PassageEtudiant, pk=passage_id, session__professeur=request.user,
    )
    sid = passage.session.id
    passage.delete()
    messages.success(request, 'Passage supprime.')
    # Redirige vers la page de planification (pas detail_session)
    return redirect('sessions_app:planifier_passages', session_id=sid)


@professeur_required
def suivi_live_passage(request, passage_id):
    """Page de suivi live : le prof voit la soutenance + notes IA en direct."""
    passage = get_object_or_404(
        PassageEtudiant, pk=passage_id, session__professeur=request.user,
    )
    return render(request, 'sessions_app/suivi_live.html', {
        'passage': passage, 'session': passage.session,
    })


# ── Code d'accès & QR ────────────────────────────────────────────────────────

@professeur_required
def code_session(request, session_id):
    """Page affichant le code, le lien et le QR code de la session."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    join_url = request.build_absolute_uri(f'/sessions/rejoindre/{session.code_acces}/')
    return render(request, 'sessions_app/code_session.html', {
        'session': session,
        'join_url': join_url,
    })


# ── Rejoindre une session (vue publique) ─────────────────────────────────────

def rejoindre_session(request, code):
    """Page publique : un étudiant rejoint une session via son code.
    Si la session appartient à une Classe, redirige vers le flux rejoindre_classe.
    """
    session = get_object_or_404(Session, code_acces=code.upper())

    # Nouveau flux : sessions rattachées à une Classe → utiliser rejoindre_classe
    if session.classe:
        return redirect('rejoindre_classe', code=session.classe.code_acces)

    # Ancien flux (sessions sans Classe) ─────────────────────────────────────
    if request.user.is_authenticated:
        if request.user.is_etudiant:
            session.inscrits.add(request.user)
            messages.success(request, f'Vous avez rejoint la session « {session.titre} ».')
            return redirect('presentation:dashboard_etudiant')
        messages.info(request, 'Seuls les étudiants peuvent rejoindre une session.')
        return redirect('accounts:dashboard_redirect')

    form = RejoindreSessionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        data = form.cleaned_data

        # Vérifier doublon email
        if data.get('email') and User.objects.filter(email=data['email']).exists():
            form.add_error('email', 'Cet email est déjà utilisé. Connectez-vous plutôt.')
        else:
            username = _norm_username(data['prenom'], data['nom'])
            user = User.objects.create_user(
                username=username,
                first_name=data['prenom'],
                last_name=data['nom'],
                email=data.get('email', ''),
                password=data['password'],
                role=User.ROLE_ETUDIANT,
                cree_par=session.professeur,
            )
            session.inscrits.add(user)
            auth_login(request, user)
            messages.success(
                request,
                f'Bienvenue {user.get_full_name()} ! Vous avez rejoint « {session.titre} ».',
            )
            return redirect('presentation:dashboard_etudiant')

    return render(request, 'sessions_app/rejoindre.html', {
        'session': session,
        'form': form,
        'code': code.upper(),
    })


# ── Planning des passages ─────────────────────────────────────────────────────

@professeur_required
@require_http_methods(['GET', 'POST'])
def planifier_passages(request, session_id):
    """Page unique de planification : manuel (formulaire inline) et auto (panneau collapsible)."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    passages = session.passages.prefetch_related('etudiants').order_by('ordre_passage')

    # IDs déjà planifiés
    deja_planifies_ids = set()
    for p in passages:
        deja_planifies_ids.update(p.etudiants.values_list('id', flat=True))

    # Source des inscrits : la Classe parente (si elle existe) sinon session.inscrits
    source_inscrits = (
        session.classe.inscrits.filter(role='etudiant')
        if session.classe
        else session.inscrits.filter(role='etudiant')
    )

    inscrits_restants = source_inscrits.exclude(
        id__in=deja_planifies_ids
    ).order_by('last_name', 'first_name')

    # Dernier passage — pour auto-suggérer l'heure suivante dans le formulaire manuel
    dernier_passage = passages.last()
    dernier_passage_iso = (
        dernier_passage.heure_prevue.strftime('%Y-%m-%dT%H:%M')
        if dernier_passage else ''
    )
    dernier_ordre = passages.aggregate(m=Max('ordre_passage'))['m'] or 0

    # Durée suggérée pour l'auto-planner = présentation + questions
    duree_suggeree = session.duree_presentation + session.duree_questions

    # Formulaires — le form manuel n'affiche que les étudiants NON encore planifiés
    form = PlanifierPassageForm(None, session=session, exclude_ids=deja_planifies_ids)
    form_auto = PlanifierAutoForm(initial={'duree_minutes': duree_suggeree, 'pause_minutes': 5})

    if request.method == 'POST':
        action = request.POST.get('action', 'nouveau_passage')

        # ── Action 1 : créer un passage manuel ──────────────────────────────
        if action == 'nouveau_passage':
            form = PlanifierPassageForm(request.POST, session=session, exclude_ids=deja_planifies_ids)
            if form.is_valid():
                etudiants_sel = list(form.cleaned_data['etudiants'])
                n = len(etudiants_sel)
                type_groupe = 'monome' if n == 1 else ('binome' if n == 2 else 'groupe')
                with transaction.atomic():
                    passage = PassageEtudiant.objects.create(
                        session=session,
                        type_groupe=type_groupe,
                        ordre_passage=form.cleaned_data['ordre_passage'],
                        heure_prevue=form.cleaned_data['heure_prevue'],
                    )
                    passage.etudiants.set(etudiants_sel)
                noms = ', '.join(u.get_full_name() or u.username for u in etudiants_sel)
                messages.success(
                    request,
                    f'Passage #{form.cleaned_data["ordre_passage"]} planifie — {noms}.',
                )
                return redirect('sessions_app:planifier_passages', session_id=session.id)

        # ── Action 2 : auto-planifier ────────────────────────────────────────
        elif action == 'auto_planifier':
            form_auto = PlanifierAutoForm(request.POST)
            if form_auto.is_valid():
                data = form_auto.cleaned_data
                etudiants = list(inscrits_restants.order_by('last_name', 'first_name'))
                if data['ordre'] == 'random':
                    random.shuffle(etudiants)

                # Taille des groupes selon le type choisi
                step = {'monome': 1, 'binome': 2, 'groupe': 3}.get(data['type_groupe'], 1)
                duree = int(data['duree_minutes'])
                pause = int(data['pause_minutes'])
                dt_debut = datetime.combine(data['date_passage'], data['heure_debut'])

                with transaction.atomic():
                    for j, i in enumerate(range(0, len(etudiants), step)):
                        groupe = etudiants[i:i + step]
                        heure_prevue = timezone.make_aware(
                            dt_debut + timedelta(minutes=j * (duree + pause))
                        )
                        passage = PassageEtudiant.objects.create(
                            session=session,
                            type_groupe=data['type_groupe'],
                            ordre_passage=dernier_ordre + j + 1,
                            heure_prevue=heure_prevue,
                        )
                        passage.etudiants.set(groupe)

                nb_passes = len(range(0, len(etudiants), step))
                messages.success(
                    request,
                    f'{nb_passes} passage(s) genere(s) automatiquement pour'
                    f' {len(etudiants)} etudiant(s).',
                )
                return redirect('sessions_app:planifier_passages', session_id=session.id)

    return render(request, 'sessions_app/planifier_passages.html', {
        'session': session,
        'classe': session.classe,
        'form': form,
        'form_auto': form_auto,
        'passages': passages,
        'inscrits_non_planifies': inscrits_restants,
        'tous_inscrits': source_inscrits.order_by('last_name', 'first_name'),
        'deja_planifies_ids': deja_planifies_ids,
        'prochain_ordre': dernier_ordre + 1,
        'duree_suggeree': duree_suggeree,
        'dernier_passage_iso': dernier_passage_iso,
        'nb_inscrits_restants': inscrits_restants.count(),
    })


@professeur_required
@require_http_methods(['POST'])
def retirer_inscrit(request, session_id, user_id):
    """Retire un étudiant de la liste des inscrits (sans supprimer son compte)."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    user = get_object_or_404(User, pk=user_id)
    session.inscrits.remove(user)
    messages.success(request, f'{user.get_full_name() or user.username} retiré(e) de la session.')
    return redirect('sessions_app:planifier_passages', session_id=session.id)


# ── Auto-planification ────────────────────────────────────────────────────────

@professeur_required
def planifier_auto(request, session_id):
    """Redirige vers la page unifiée de planification (panneau auto intégré dedans)."""
    return redirect('sessions_app:planifier_passages', session_id=session_id)


@professeur_required
@require_http_methods(['POST'])
def recalculer_planning(request, session_id):
    """Recalcule les heure_prevue de tous les passages en séquence à partir d'une heure de départ."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    heure_ref_str = request.POST.get('heure_ref', '')
    try:
        pause = max(0, int(request.POST.get('pause_minutes', 0)))
    except (ValueError, TypeError):
        pause = 0

    duree_slot = session.duree_presentation + session.duree_questions

    try:
        dt_ref = datetime.strptime(heure_ref_str, '%Y-%m-%dT%H:%M')
        dt_ref = timezone.make_aware(dt_ref)
    except (ValueError, TypeError):
        messages.error(request, 'Heure de reference invalide.')
        return redirect('sessions_app:planifier_passages', session_id=session_id)

    passages = list(session.passages.order_by('ordre_passage'))
    if not passages:
        messages.warning(request, 'Aucun passage a recalculer.')
        return redirect('sessions_app:planifier_passages', session_id=session_id)

    with transaction.atomic():
        for j, p in enumerate(passages):
            p.heure_prevue = dt_ref + timedelta(minutes=j * (duree_slot + pause))
            p.save(update_fields=['heure_prevue'])

    messages.success(
        request,
        f'Planning recalcule : {len(passages)} passage(s) mis a jour'
        f' (creneau = {duree_slot} min, pause = {pause} min).',
    )
    return redirect('sessions_app:planifier_passages', session_id=session_id)


# ── Import fichier (CSV/XLSX/PDF/image) ──────────────────────────────────────

@professeur_required
def importer_groupes(request, session_id):
    """Étape 1 : upload du fichier + prévisualisation du matching."""
    from django.conf import settings

    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    classe = session.classe

    if classe:
        inscrits = classe.inscrits.filter(role='etudiant')
    else:
        inscrits = session.inscrits.filter(role='etudiant')

    form = ImportGroupesForm(request.POST or None, request.FILES or None)
    preview = None
    erreur = None

    if request.method == 'POST' and form.is_valid():
        fichier = request.FILES['fichier']
        content = fichier.read()
        filename = fichier.name

        try:
            import anthropic as _anthropic
            api_key = getattr(settings, 'ANTHROPIC_API_KEY', '')
            lignes = parse_fichier(content, filename, api_key=api_key)
            if not lignes:
                erreur = 'Aucun étudiant trouvé dans le fichier.'
            else:
                groupes = matcher_groupes(lignes, inscrits)
                preview = groupes_to_json(groupes)
                # Stocker en session Django pour la confirmation
                request.session['import_preview'] = {
                    'session_id': session_id,
                    'groupes': preview,
                    'heure_debut': form.cleaned_data['heure_debut'].strftime('%Y-%m-%dT%H:%M'),
                    'pause_minutes': form.cleaned_data['pause_minutes'],
                }
        except _anthropic.AuthenticationError:
            erreur = (
                'Clé API Anthropic invalide ou expirée (erreur 401). '
                'Vérifiez ANTHROPIC_API_KEY dans votre fichier .env.'
            )
        except _anthropic.APIConnectionError:
            erreur = (
                'Impossible de joindre l\'API Anthropic (erreur réseau). '
                'Vérifiez votre connexion Internet et réessayez.'
            )
        except _anthropic.RateLimitError:
            erreur = (
                'Limite de débit Anthropic atteinte. '
                'Attendez quelques secondes et réessayez.'
            )
        except Exception as exc:
            erreur = str(exc)

    return render(request, 'sessions_app/importer_groupes.html', {
        'session': session,
        'classe': classe,
        'form': form,
        'preview': preview,
        'erreur': erreur,
        'duree_suggeree': session.duree_presentation + session.duree_questions,
    })


@professeur_required
@require_http_methods(['POST'])
def importer_groupes_confirmer(request, session_id):
    """Étape 2 : création des passages à partir des groupes validés."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    data = request.session.pop('import_preview', None)

    if not data or data.get('session_id') != session_id:
        messages.error(request, 'Session d\'import expirée. Veuillez recommencer.')
        return redirect('sessions_app:importer_groupes', session_id=session_id)

    groupes_valides = [g for g in data['groupes'] if g['valide']]
    if not groupes_valides:
        messages.warning(request, 'Aucun groupe valide à importer.')
        return redirect('sessions_app:importer_groupes', session_id=session_id)

    try:
        dt_debut = timezone.make_aware(
            datetime.strptime(data['heure_debut'], '%Y-%m-%dT%H:%M')
        )
    except (ValueError, TypeError):
        messages.error(request, 'Heure de début invalide.')
        return redirect('sessions_app:importer_groupes', session_id=session_id)

    pause = int(data.get('pause_minutes', 5))
    duree_slot = session.duree_presentation + session.duree_questions

    dernier_ordre = session.passages.aggregate(m=Max('ordre_passage'))['m'] or 0

    nb_crees = 0
    with transaction.atomic():
        for j, groupe in enumerate(groupes_valides):
            user_ids = [m['user_id'] for m in groupe['membres'] if m['user_id']]
            if not user_ids:
                continue
            heure_prevue = dt_debut + timedelta(minutes=j * (duree_slot + pause))
            passage = PassageEtudiant.objects.create(
                session=session,
                type_groupe=groupe['type_groupe'],
                ordre_passage=dernier_ordre + j + 1,
                heure_prevue=heure_prevue,
            )
            passage.etudiants.set(User.objects.filter(id__in=user_ids))
            nb_crees += 1

    messages.success(
        request,
        f'{nb_crees} passage(s) créé(s) depuis le fichier importé.'
    )
    return redirect('sessions_app:planifier_passages', session_id=session_id)


# ── Export planning XLSX ──────────────────────────────────────────────────────

@professeur_required
def exporter_planning_xlsx(request, session_id):
    """Génère et télécharge le planning complet au format XLSX."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    passages = session.passages.order_by('ordre_passage').prefetch_related('etudiants')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Planning'

    # ── Styles ──────────────────────────────────────────────────────────────
    couleur_entete  = '1e3a5f'   # bleu foncé
    couleur_pair    = 'eef2ff'   # bleu très clair
    couleur_impair  = 'ffffff'   # blanc
    couleur_bord    = 'c7d2fe'

    police_titre    = Font(name='Calibri', size=14, bold=True, color='1e3a5f')
    police_entete   = Font(name='Calibri', size=10, bold=True, color='ffffff')
    police_normale  = Font(name='Calibri', size=10)
    police_heure    = Font(name='Consolas', size=10, bold=True, color='1d4ed8')

    remplissage_entete = PatternFill('solid', fgColor=couleur_entete)
    remplissage_pair   = PatternFill('solid', fgColor=couleur_pair)
    remplissage_impair = PatternFill('solid', fgColor=couleur_impair)

    bord_fin = Border(
        left=Side(style='thin', color=couleur_bord),
        right=Side(style='thin', color=couleur_bord),
        top=Side(style='thin', color=couleur_bord),
        bottom=Side(style='thin', color=couleur_bord),
    )
    centre  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gauche  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Titre (ligne 1) ──────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    cell_titre = ws['A1']
    cell_titre.value = f'Planning des passages — {session.titre}'
    cell_titre.font = police_titre
    cell_titre.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Sous-titre ligne 2
    ws.merge_cells('A2:F2')
    cell_sub = ws['A2']
    duree_slot = session.duree_presentation + session.duree_questions
    cell_sub.value = (
        f'Durée par créneau : {duree_slot} min '
        f'({session.duree_presentation} min présentation + {session.duree_questions} min questions)  '
        f'| Total : {passages.count()} passage(s)'
    )
    cell_sub.font = Font(name='Calibri', size=9, italic=True, color='64748b')
    cell_sub.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── En-têtes (ligne 3) ──────────────────────────────────────────────────
    ENTETES = ['#', 'Étudiant(s)', 'Type', 'Heure prévue', 'Statut', 'Notes']
    for col_idx, label in enumerate(ENTETES, 1):
        c = ws.cell(row=3, column=col_idx, value=label)
        c.font = remplissage_entete and police_entete
        c.fill = remplissage_entete
        c.alignment = centre
        c.border = bord_fin
    ws.row_dimensions[3].height = 20

    # ── Données ──────────────────────────────────────────────────────────────
    STATUT_FR = {
        'en_attente': 'En attente',
        'en_cours':   'En cours',
        'questions':  'Questions',
        'termine':    'Terminé',
        'note':       'Noté',
    }

    for idx, p in enumerate(passages, 1):
        row = idx + 3   # données à partir de la ligne 4
        remplissage = remplissage_pair if idx % 2 == 0 else remplissage_impair

        noms = ', '.join(
            e.get_full_name() or e.username
            for e in p.etudiants.all()
        )
        heure = p.heure_prevue.strftime('%d/%m/%Y  %H:%M') if p.heure_prevue else '—'
        statut = STATUT_FR.get(p.statut, p.statut)

        TYPE_FR = {
            'monome': 'Monôme (1)',
            'binome': 'Binôme (2)',
            'groupe': f'Groupe ({p.etudiants.count()})',
        }
        type_label = TYPE_FR.get(p.type_groupe, p.type_groupe)

        valeurs = [p.ordre_passage, noms, type_label, heure, statut, '']
        for col_idx, val in enumerate(valeurs, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = police_heure if col_idx == 4 else police_normale
            c.fill = remplissage
            c.border = bord_fin
            c.alignment = centre if col_idx in (1, 3, 4, 5) else gauche

        ws.row_dimensions[row].height = 22

    # ── Largeurs des colonnes ────────────────────────────────────────────────
    largeurs = [6, 55, 18, 20, 14, 25]
    for col_idx, larg in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larg

    # Figer les en-têtes
    ws.freeze_panes = 'A4'

    # ── Réponse HTTP ─────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nom_fichier = (
        'planning_'
        + session.titre[:30].replace(' ', '_').replace('/', '-')
        + '.xlsx'
    )
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response


# ── Notes de la session (tableau récap + export XLSX) ─────────────────────────

def _construire_notes_session(session):
    """Retourne (criteres, lignes) pour l'affichage ou l'export des notes.

    criteres : liste ordonnée de CritereNotation
    lignes   : liste de dicts {passage, etudiant, notes_par_critere, globale}
                 notes_par_critere = {critere_id: NoteIA ou None}
    """
    from notation.models import NoteIA

    criteres = list(session.criteres.all().order_by('ordre', 'id'))
    passages = session.passages.prefetch_related('etudiants').order_by('ordre_passage')

    lignes = []
    for passage in passages:
        for etudiant in passage.etudiants.all():
            notes_qs = NoteIA.objects.filter(
                passage=passage, etudiant=etudiant
            ).select_related('critere')
            notes_map = {n.critere_id: n for n in notes_qs}

            total_p, total_c = 0.0, 0.0
            for c in criteres:
                n = notes_map.get(c.id)
                if n is not None:
                    total_p += n.note_finale * c.coefficient
                    total_c += c.coefficient
            globale = round(total_p / total_c, 2) if total_c else None

            lignes.append({
                'passage': passage,
                'etudiant': etudiant,
                'notes_par_critere': notes_map,
                'globale': globale,
            })

    return criteres, lignes


@professeur_required
def notes_session(request, session_id):
    """Tableau récapitulatif de toutes les notes pour une session."""
    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    criteres, lignes = _construire_notes_session(session)
    return render(request, 'sessions_app/notes_session.html', {
        'session': session,
        'criteres': criteres,
        'lignes': lignes,
    })


@professeur_required
def exporter_notes_xlsx(request, session_id):
    """Génère et télécharge les notes de tous les étudiants au format XLSX."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    session = get_object_or_404(Session, pk=session_id, professeur=request.user)
    criteres, lignes = _construire_notes_session(session)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notes'

    # ── Styles ───────────────────────────────────────────────────────────────
    police_titre   = Font(name='Calibri', size=13, bold=True, color='1e3a5f')
    police_entete  = Font(name='Calibri', size=10, bold=True, color='ffffff')
    police_note    = Font(name='Consolas', size=10, bold=True)
    police_norm    = Font(name='Calibri', size=10)
    remplissage_h  = PatternFill('solid', fgColor='1e3a5f')
    remplissage_p  = PatternFill('solid', fgColor='eef2ff')
    remplissage_i  = PatternFill('solid', fgColor='ffffff')
    bord = Border(
        left=Side(style='thin', color='c7d2fe'),
        right=Side(style='thin', color='c7d2fe'),
        top=Side(style='thin', color='c7d2fe'),
        bottom=Side(style='thin', color='c7d2fe'),
    )
    centre = Alignment(horizontal='center', vertical='center')
    gauche = Alignment(horizontal='left', vertical='center')

    # ── Titre ────────────────────────────────────────────────────────────────
    nb_cols = 3 + len(criteres) + 1  # #, Nom, Prénom + critères + Moyenne
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws['A1'].value = f'Notes — {session.titre}'
    ws['A1'].font  = police_titre
    ws['A1'].alignment = gauche
    ws.row_dimensions[1].height = 22

    # ── En-têtes ──────────────────────────────────────────────────────────────
    en_tetes = ['#', 'Nom', 'Prénom'] + [c.nom for c in criteres] + ['Moyenne /20']
    for col, label in enumerate(en_tetes, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font  = police_entete
        c.fill  = remplissage_h
        c.border = bord
        c.alignment = centre
    ws.row_dimensions[2].height = 22

    # ── Données ──────────────────────────────────────────────────────────────
    for idx, ligne in enumerate(lignes, 1):
        row = idx + 2
        remplissage = remplissage_p if idx % 2 == 0 else remplissage_i
        etudiant = ligne['etudiant']
        notes_map = ligne['notes_par_critere']

        valeurs = [
            ligne['passage'].ordre_passage,
            etudiant.last_name,
            etudiant.first_name,
        ]
        for c in criteres:
            n = notes_map.get(c.id)
            valeurs.append(round(n.note_finale, 1) if n else '')

        globale = ligne['globale']
        valeurs.append(round(globale, 2) if globale is not None else '')

        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = remplissage
            cell.border = bord
            is_note = col > 3
            cell.font      = police_note if is_note else police_norm
            cell.alignment = centre if is_note else gauche
            # Mise en couleur de la moyenne
            if col == nb_cols and isinstance(val, (int, float)):
                if val >= 14:
                    cell.font = Font(name='Consolas', size=10, bold=True, color='16a34a')
                elif val >= 10:
                    cell.font = Font(name='Consolas', size=10, bold=True, color='d97706')
                else:
                    cell.font = Font(name='Consolas', size=10, bold=True, color='dc2626')
        ws.row_dimensions[row].height = 18

    # ── Largeurs ─────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    for i in range(len(criteres)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 16
    ws.column_dimensions[get_column_letter(nb_cols)].width = 14

    ws.freeze_panes = 'A3'

    # ── Réponse ──────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nom = 'notes_' + session.titre[:30].replace(' ', '_').replace('/', '-') + '.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response
